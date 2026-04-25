from __future__ import annotations

import json
import re
import subprocess
from collections import OrderedDict, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_XLSX_CANDIDATES = [
    ROOT / "hypnocards-v2-he" / "דפוסי שפה עברית" / "קלפים טבלה ארוכה (1).xlsx",
    ROOT.parent.parent / "דפוסי שפה עברית" / "קלפים טבלה ארוכה (1).xlsx",
]
SOURCE_HTML_CANDIDATES = [
    ROOT / "hypnocards-v2-he" / "דפוסי שפה עברית" / "index.html",
    ROOT.parent.parent / "דפוסי שפה עברית" / "index.html",
]
OUT_MANIFEST = ROOT / "hypnocards-v2-he" / "trance-he" / "trance-he-manifest.json"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"\s+", " ", text)


def resolve_first_existing(candidates: list[Path]) -> Path | None:
    for p in candidates:
        if p.exists():
            return p
    return None


def load_xlsx_rows() -> list[tuple]:
    source_xlsx = resolve_first_existing(SOURCE_XLSX_CANDIDATES)
    if not source_xlsx:
        raise FileNotFoundError(f"missing source xlsx in candidates: {SOURCE_XLSX_CANDIDATES}")
    wb = load_workbook(source_xlsx, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        rows.append(row)
    return rows


def build_cards(rows: list[tuple]) -> list[dict]:
    grouped: dict[str, dict] = {}
    discovered_order: dict[str, int] = {}
    ignored_themes = {"FullExample", "full example", "nan", "none"}
    for r in rows:
        # xlsx columns: ID, Category, Pattern, Theme, Content, Color
        card_id = normalize_text(r[0])
        category = normalize_text(r[1])
        pattern = normalize_text(r[2])
        theme = normalize_text(r[3])
        content = normalize_text(r[4])
        color = normalize_text(r[5]) or "#9CA3AF"

        if not card_id or not category or not pattern or not theme or not content:
            continue
        if theme in ignored_themes:
            continue

        if card_id not in grouped:
            grouped[card_id] = {
                "id": card_id,
                "category": category,
                "pattern": pattern,
                "color": color,
                "variations": defaultdict(list),
                "_order_hint": 10**9,
            }
        if category and category not in discovered_order:
            discovered_order[category] = len(discovered_order)

        card = grouped[card_id]
        # Keep first stable values in case of inconsistent duplicate rows.
        if not card["category"]:
            card["category"] = category
        if not card["pattern"]:
            card["pattern"] = pattern
        if not card["color"] or card["color"] == "#9CA3AF":
            card["color"] = color

        if category in discovered_order:
            card["_order_hint"] = min(card["_order_hint"], discovered_order[category] * 1000)

        vals = card["variations"][theme]
        if content not in vals:
            vals.append(content)

    built = []
    for card in grouped.values():
        variations = OrderedDict()
        for theme, texts in card["variations"].items():
            if texts:
                variations[theme] = texts
        if not variations:
            continue
        built.append(
            {
                "id": card["id"],
                "pattern": card["pattern"],
                "category": card["category"],
                "color": card["color"],
                "variations": variations,
                "_order_hint": card["_order_hint"],
            }
        )

    # Match source order as closely as possible (embedded order then numeric ID fallback)
    def sort_key(c: dict) -> tuple[int, int]:
        hint = int(c.get("_order_hint", 10**9))
        try:
            id_num = int(str(c.get("id", "0")))
        except ValueError:
            id_num = 0
        return (hint, id_num)

    built.sort(key=sort_key)
    for c in built:
        c.pop("_order_hint", None)
    return built


def build_categories(cards: list[dict]) -> list[dict]:
    color_by_cat: OrderedDict[str, str] = OrderedDict()
    for c in cards:
        category = normalize_text(c.get("category"))
        color = normalize_text(c.get("color")) or "#9CA3AF"
        if category and category not in color_by_cat:
            color_by_cat[category] = color
    return [{"name": name, "color": color} for name, color in color_by_cat.items()]


def extract_legacy_meta() -> dict:
    source_html = resolve_first_existing(SOURCE_HTML_CANDIDATES)
    if not source_html:
        return {"cards": [], "categories": [], "category_info": {}}
    js = r"""
const fs = require('fs');
const htmlPath = process.argv[1];
const html = fs.readFileSync(htmlPath, 'utf8');
const marker = 'const EMBEDDED_DATA = JSON.parse(`';
const start = html.indexOf(marker);
if (start < 0) {
  process.stdout.write(JSON.stringify({ cards: [], categories: [], category_info: {} }));
  process.exit(0);
}
const payloadStart = start + marker.length;
const end = html.indexOf('`);', payloadStart);
if (end < 0) {
  process.stdout.write(JSON.stringify({ cards: [], categories: [], category_info: {} }));
  process.exit(0);
}
const payload = html.slice(payloadStart, end);
let jsonText = payload;
if (jsonText.startsWith('`')) jsonText = jsonText.slice(1);
if (jsonText.endsWith('`')) jsonText = jsonText.slice(0, -1);
const decoded = Function('"use strict"; return `' + jsonText.replace(/`/g, '\\`') + '`;')();
const data = JSON.parse(decoded);
process.stdout.write(JSON.stringify({
  cards: Array.isArray(data.cards) ? data.cards : [],
  categories: Array.isArray(data.categories) ? data.categories : [],
  category_info: (data.category_info && typeof data.category_info === 'object') ? data.category_info : {}
}));
"""
    out = subprocess.check_output(
        ["node", "-e", js, str(source_html)],
        text=True,
        encoding="utf-8",
    )
    try:
        parsed = json.loads(out)
    except json.JSONDecodeError:
        return {"cards": [], "categories": [], "category_info": {}}
    if not isinstance(parsed, dict):
        return {"cards": [], "categories": [], "category_info": {}}
    return {
        "cards": parsed.get("cards", []) if isinstance(parsed.get("cards", []), list) else [],
        "categories": parsed.get("categories", []) if isinstance(parsed.get("categories", []), list) else [],
        "category_info": parsed.get("category_info", {}) if isinstance(parsed.get("category_info", {}), dict) else {},
    }


def validate(cards: list[dict], categories: list[dict]) -> None:
    ids = [str(c.get("id", "")).strip() for c in cards]
    if len(ids) != len(set(ids)):
        dup_ids = sorted({x for x in ids if ids.count(x) > 1})
        raise RuntimeError(f"duplicate card ids found: {dup_ids[:10]}")
    if not cards:
        raise RuntimeError("no cards built for trance manifest")
    if not categories:
        raise RuntimeError("no categories present for trance manifest")


def main() -> None:
    legacy_meta = extract_legacy_meta()
    legacy_cards = legacy_meta.get("cards", [])
    legacy_categories = legacy_meta.get("categories", [])
    legacy_category_info = legacy_meta.get("category_info", {})
    if legacy_cards:
        cards = legacy_cards
        categories = legacy_categories if legacy_categories else build_categories(cards)
    else:
        rows = load_xlsx_rows()
        cards = build_cards(rows)
        categories = legacy_categories if legacy_categories else build_categories(cards)

    manifest = {
        "cards": cards,
        "categories": categories,
        "category_info": legacy_category_info,
    }
    validate(manifest["cards"], manifest["categories"])

    OUT_MANIFEST.parent.mkdir(parents=True, exist_ok=True)
    OUT_MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    print("manifest_generated:", True)
    print("cards:", len(manifest["cards"]))
    print("categories:", len(manifest["categories"]))


if __name__ == "__main__":
    main()
